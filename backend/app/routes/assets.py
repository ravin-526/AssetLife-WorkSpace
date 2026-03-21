from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from bson import ObjectId
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.security import get_current_user
from app.core.status_master import (
    get_default_status_name,
    get_status_lookup,
    resolve_status_from_lookup,
    validate_or_map_status,
)
from app.db.mongo import get_db
from app.routes.categories import FINAL_CATEGORY_SUBCATEGORIES

router = APIRouter(prefix="/api/assets", tags=["Assets"])

ASSET_SOURCE_EMAIL_SYNC = "email_sync"
ASSET_SOURCE_INVOICE_UPLOAD = "invoice_upload"
ASSET_SOURCE_EXCEL_UPLOAD = "excel_upload"
ASSET_SOURCE_QR_SCAN = "qr_scan"
ASSET_SOURCE_MANUAL = "manual"

ASSET_SOURCE_ALIASES = {
    ASSET_SOURCE_EMAIL_SYNC: ASSET_SOURCE_EMAIL_SYNC,
    "gmail": ASSET_SOURCE_EMAIL_SYNC,
    "email": ASSET_SOURCE_EMAIL_SYNC,
    ASSET_SOURCE_INVOICE_UPLOAD: ASSET_SOURCE_INVOICE_UPLOAD,
    "invoice": ASSET_SOURCE_INVOICE_UPLOAD,
    ASSET_SOURCE_EXCEL_UPLOAD: ASSET_SOURCE_EXCEL_UPLOAD,
    "excel": ASSET_SOURCE_EXCEL_UPLOAD,
    ASSET_SOURCE_QR_SCAN: ASSET_SOURCE_QR_SCAN,
    "barcode_qr": ASSET_SOURCE_QR_SCAN,
    "qr": ASSET_SOURCE_QR_SCAN,
    "barcode": ASSET_SOURCE_QR_SCAN,
    ASSET_SOURCE_MANUAL: ASSET_SOURCE_MANUAL,
    "manual_entry": ASSET_SOURCE_MANUAL,
}

UPLOAD_ROOT = Path(__file__).resolve().parents[2] / "uploads"
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
DOC_ROOT = UPLOAD_ROOT / "documents"
DOC_ROOT.mkdir(parents=True, exist_ok=True)

EXCEL_TEMPLATE_FIELDS = [
    {"key": "product_name", "label": "Asset Name *", "type": "text", "required": True},
    {"key": "category", "label": "Category *", "type": "category", "required": True},
    {"key": "custom_category", "label": "Enter Category (if Category = Other)", "type": "text", "required": False},
    {"key": "subcategory", "label": "SubCategory *", "type": "subcategory", "required": True},
    {"key": "custom_subcategory", "label": "Enter SubCategory (if SubCategory = Other)", "type": "text", "required": False},
    {"key": "vendor", "label": "Vendor", "type": "text", "required": False},
    {"key": "purchase_date", "label": "Purchase Date", "type": "date", "required": False},
    {"key": "price", "label": "Purchase Price", "type": "number", "required": False},
    {"key": "serial_number", "label": "Serial Number", "type": "text", "required": False},
    {"key": "model_number", "label": "Model Number", "type": "text", "required": False},
    {"key": "invoice_number", "label": "Invoice Number", "type": "text", "required": False},
    {"key": "location", "label": "Location", "type": "text", "required": False},
    {"key": "assigned_user", "label": "Assigned User", "type": "text", "required": False},
    {"key": "description", "label": "Description", "type": "text", "required": False},
    {"key": "notes", "label": "Notes", "type": "text", "required": False},
    {"key": "warranty_available", "label": "Warranty Available", "type": "boolean", "required": False},
    {"key": "warranty_provider", "label": "Warranty Provider", "type": "text", "required": False},
    {"key": "warranty_type", "label": "Warranty Type", "type": "list", "required": False},
    {"key": "warranty_start_date", "label": "Warranty Start Date", "type": "date", "required": False},
    {"key": "warranty_end_date", "label": "Warranty End Date", "type": "date", "required": False},
    {"key": "warranty_notes", "label": "Warranty Notes", "type": "text", "required": False},
    {"key": "warranty_reminder_30_days", "label": "Warranty Reminder 30 Days", "type": "boolean", "required": False},
    {"key": "warranty_reminder_7_days", "label": "Warranty Reminder 7 Days", "type": "boolean", "required": False},
    {"key": "warranty_reminder_on_expiry", "label": "Warranty Reminder On Expiry", "type": "boolean", "required": False},
    {"key": "insurance_available", "label": "Insurance Available", "type": "boolean", "required": False},
    {"key": "insurance_provider", "label": "Insurance Provider", "type": "text", "required": False},
    {"key": "insurance_policy_number", "label": "Policy Number", "type": "text", "required": False},
    {"key": "insurance_start_date", "label": "Insurance Start Date", "type": "date", "required": False},
    {"key": "insurance_expiry_date", "label": "Insurance Expiry Date", "type": "date", "required": False},
    {"key": "insurance_premium_amount", "label": "Insurance Premium Amount", "type": "number", "required": False},
    {"key": "insurance_coverage_notes", "label": "Insurance Notes", "type": "text", "required": False},
    {"key": "insurance_reminder_45_days", "label": "Insurance Reminder 45 Days", "type": "boolean", "required": False},
    {"key": "insurance_reminder_15_days", "label": "Insurance Reminder 15 Days", "type": "boolean", "required": False},
    {"key": "service_required", "label": "Service Required", "type": "boolean", "required": False},
    {"key": "service_frequency", "label": "Service Frequency", "type": "list", "required": False},
    {"key": "service_custom_interval_days", "label": "Service Interval (Days)", "type": "number", "required": False},
    {"key": "service_reminder_enabled", "label": "Enable Next Service Reminder", "type": "boolean", "required": False},
]

EXCEL_TEMPLATE_COLUMNS = [field["key"] for field in EXCEL_TEMPLATE_FIELDS]
EXCEL_TEMPLATE_HEADERS = [field["label"] for field in EXCEL_TEMPLATE_FIELDS]
EXCEL_TEMPLATE_HEADER_BY_KEY = {field["key"]: field["label"] for field in EXCEL_TEMPLATE_FIELDS}
EXCEL_TEMPLATE_FIELD_BY_KEY = {field["key"]: field for field in EXCEL_TEMPLATE_FIELDS}
BOOLEAN_TEMPLATE_VALUES = ["Yes", "No"]
WARRANTY_TYPE_VALUES = ["manufacturer", "extended"]
SERVICE_FREQUENCY_VALUES = ["monthly", "quarterly", "half_yearly", "yearly", "custom"]
EXCEL_TEMPLATE_ROW_LIMIT = 2000
EXCEL_TEMPLATE_SHEET_NAME = "Assets"
EXCEL_TEMPLATE_MASTER_SHEET_NAME = "MasterData"
EXCEL_TEMPLATE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="DCEBFF")
EXCEL_TEMPLATE_HEADER_FONT = Font(bold=True, color="1F2937")


def _sample_asset_row(**overrides: Any) -> dict[str, Any]:
    row = {
        "product_name": "",
        "category": "Electronics",
        "custom_category": "",
        "subcategory": "Laptops",
        "custom_subcategory": "",
        "vendor": "Croma",
        "purchase_date": "2024-01-01",
        "price": 0,
        "serial_number": "",
        "model_number": "",
        "invoice_number": "",
        "location": "Bangalore HQ",
        "assigned_user": "Operations Team",
        "description": "",
        "notes": "",
        "warranty_available": "Yes",
        "warranty_provider": "Manufacturer Warranty",
        "warranty_type": "manufacturer",
        "warranty_start_date": "2024-01-01",
        "warranty_end_date": "2025-01-01",
        "warranty_notes": "Standard coverage",
        "warranty_reminder_30_days": "Yes",
        "warranty_reminder_7_days": "Yes",
        "warranty_reminder_on_expiry": "Yes",
        "insurance_available": "No",
        "insurance_provider": "",
        "insurance_policy_number": "",
        "insurance_start_date": "",
        "insurance_expiry_date": "",
        "insurance_premium_amount": "",
        "insurance_coverage_notes": "",
        "insurance_reminder_45_days": "No",
        "insurance_reminder_15_days": "No",
        "service_required": "Yes",
        "service_frequency": "yearly",
        "service_custom_interval_days": "",
        "service_reminder_enabled": "Yes",
    }
    row.update(overrides)
    return row


EXCEL_TEMPLATE_SAMPLE_ROWS = [
    _sample_asset_row(product_name="MacBook Pro 16", category="Electronics", subcategory="Laptops", vendor="iStore", purchase_date="2024-01-12", price=249999, serial_number="MBP16-001", model_number="A2780", invoice_number="INV-APL-2024-001", location="Bangalore HQ", assigned_user="Asha N", description="Engineering laptop", notes="Issued to design team", warranty_provider="Apple Care", warranty_end_date="2027-01-11", warranty_notes="3 year plan", insurance_available="Yes", insurance_provider="HDFC Ergo", insurance_policy_number="POL-APL-9910", insurance_start_date="2024-01-12", insurance_expiry_date="2025-01-11", insurance_premium_amount=12999, insurance_coverage_notes="Accidental damage cover"),
    _sample_asset_row(product_name="Dell XPS 13", category="Electronics", subcategory="Laptops", vendor="Dell Store", purchase_date="2024-02-03", price=145000, serial_number="DX13-7781", model_number="XPS-9315", invoice_number="INV-DEL-2024-013", location="Mumbai Office", assigned_user="Rohit M", description="Sales leadership laptop", warranty_provider="Dell Warranty", warranty_type="extended", warranty_end_date="2026-02-02", warranty_notes="Extended coverage", service_frequency="half_yearly"),
    _sample_asset_row(product_name="iPhone 15", category="Electronics", subcategory="Mobile Phones", vendor="Unicorn", purchase_date="2024-03-15", price=89900, serial_number="IPH15-9981", model_number="A3090", invoice_number="INV-APL-2024-089", location="Chennai Branch", assigned_user="Priya S", description="Executive phone", notes="Dual SIM setup", warranty_provider="Apple Care+", warranty_end_date="2026-03-14", insurance_available="Yes", insurance_provider="Bajaj Allianz", insurance_policy_number="POL-IPH-4401", insurance_start_date="2024-03-15", insurance_expiry_date="2025-03-14", insurance_premium_amount=5500, insurance_coverage_notes="Screen and theft cover"),
    _sample_asset_row(product_name="Samsung Galaxy Tab S9", category="Electronics", subcategory="Tablets", vendor="Samsung Plaza", purchase_date="2024-03-21", price=72999, serial_number="TABS9-201", model_number="SM-X710", invoice_number="INV-SAM-2024-090", location="Hyderabad Office", assigned_user="Field Team", description="Customer demo tablet", notes="Includes stylus", warranty_provider="Samsung Care", warranty_end_date="2026-03-20", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="LG 27 Monitor", category="Electronics", subcategory="Monitors", vendor="Reliance Digital", purchase_date="2024-04-02", price=22999, serial_number="LGMON-873", model_number="27UP650", invoice_number="INV-LG-2024-115", location="Pune Office", assigned_user="Design Bay", description="4K productivity monitor", warranty_provider="LG India", warranty_end_date="2027-04-01"),
    _sample_asset_row(product_name="HP LaserJet Pro", category="Electronics", subcategory="Printers & Scanners", vendor="HP Store", purchase_date="2024-04-11", price=32990, serial_number="HPLJ-2190", model_number="MFP-4104", invoice_number="INV-HP-2024-045", location="Admin Floor", assigned_user="Admin Team", description="High-volume office printer", warranty_provider="HP", warranty_end_date="2025-04-10", service_frequency="quarterly"),
    _sample_asset_row(product_name="TP-Link Archer AX55", category="Electronics", subcategory="Routers & Modems", vendor="Amazon Business", purchase_date="2024-04-18", price=8999, serial_number="TPL-AX55-91", model_number="AX55", invoice_number="INV-TPL-2024-028", location="Network Room", assigned_user="Infra Team", description="Backup office router", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="SanDisk Extreme SSD 1TB", category="Electronics", subcategory="External Storage (HDD/SSD)", vendor="Croma", purchase_date="2024-04-30", price=11499, serial_number="SDK-SSD-1001", model_number="SDSSDE61", invoice_number="INV-SDK-2024-041", location="Media Lab", assigned_user="Content Team", description="Portable backup SSD", warranty_provider="SanDisk", warranty_end_date="2029-04-29", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="OnePlus Power Bank 20000mAh", category="Electronics", subcategory="Power Banks", vendor="OnePlus Store", purchase_date="2024-05-05", price=2499, serial_number="PB-20000-19", model_number="PB20K", invoice_number="INV-OP-2024-053", location="Travel Desk", assigned_user="Sales Team", description="Travel accessory stock", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="boAt Charger 65W", category="Electronics", subcategory="Chargers & Adapters", vendor="Flipkart Business", purchase_date="2024-05-12", price=1999, serial_number="BOAT-CHR-65", model_number="Super65", invoice_number="INV-BOAT-2024-061", location="IT Store", assigned_user="Support Desk", description="Laptop fast charger", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="Samsung 55 Smart TV", category="Home Appliances", subcategory="Televisions", vendor="Reliance Digital", purchase_date="2024-05-28", price=62999, serial_number="SAMTV55-773", model_number="UA55AU7700", invoice_number="INV-SAM-2024-114", location="Meeting Room 2", assigned_user="Facilities", description="Meeting room display", notes="Wall mounted", warranty_provider="Samsung Care", warranty_type="extended", warranty_end_date="2027-05-27", insurance_available="Yes", insurance_provider="ICICI Lombard", insurance_policy_number="POL-TV-7782", insurance_start_date="2024-05-28", insurance_expiry_date="2025-05-27", insurance_premium_amount=3800, insurance_coverage_notes="Power surge cover"),
    _sample_asset_row(product_name="Daikin Split AC", category="Home Appliances", subcategory="Air Conditioners", vendor="Cooling Hub", purchase_date="2024-06-04", price=48750, serial_number="DKAC-8804", model_number="FTKF50", invoice_number="INV-DKN-2024-061", location="Server Room", assigned_user="Infra Team", description="Critical cooling unit", notes="24x7 operation", warranty_provider="Daikin", warranty_end_date="2026-06-03", insurance_available="Yes", insurance_provider="Tata AIG", insurance_policy_number="POL-AC-9012", insurance_start_date="2024-06-04", insurance_expiry_date="2025-06-03", insurance_premium_amount=2750, insurance_coverage_notes="Compressor cover", service_frequency="half_yearly"),
    _sample_asset_row(product_name="Godrej Refrigerator", category="Home Appliances", subcategory="Refrigerators", vendor="HomeTown", purchase_date="2024-06-13", price=28990, serial_number="GDRF-5501", model_number="RT-EON", invoice_number="INV-GOD-2024-033", location="Pantry", assigned_user="Office Ops", description="Office pantry refrigerator", warranty_provider="Godrej Care", warranty_end_date="2027-06-12"),
    _sample_asset_row(product_name="Voltas Water Purifier", category="Home Appliances", subcategory="Water Purifiers", vendor="Vijay Sales", purchase_date="2024-06-24", price=15499, serial_number="VT-WP-331", model_number="RO-Plus", invoice_number="INV-VOL-2024-044", location="Cafeteria", assigned_user="Facilities", description="RO water purifier", warranty_provider="Voltas", warranty_end_date="2025-06-23", service_frequency="quarterly"),
    _sample_asset_row(product_name="Blue Star Air Cooler", category="Home Appliances", subcategory="Air Coolers", vendor="Amazon Business", purchase_date="2024-07-02", price=10999, serial_number="BSTAR-ACR-88", model_number="CoolBreeze", invoice_number="INV-BS-2024-050", location="Reception", assigned_user="Front Office", description="Seasonal cooling unit", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="IKEA Ergonomic Chair", category="Furniture", subcategory="Chairs", vendor="IKEA", purchase_date="2024-07-08", price=12999, serial_number="IKEA-CHR-204", model_number="MARKUS", invoice_number="INV-IK-2024-071", location="Bangalore HQ", assigned_user="Engineering", description="Ergonomic workstation chair", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", insurance_available="No", service_frequency="", service_required="No", service_reminder_enabled="No"),
    _sample_asset_row(product_name="Conference Table 10-Seater", category="Furniture", subcategory="Tables", vendor="Urban Ladder Business", purchase_date="2024-07-14", price=45999, serial_number="CONF-TBL-10", model_number="UL-CF10", invoice_number="INV-UL-2024-081", location="Board Room", assigned_user="Admin Team", description="Solid wood conference table", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", insurance_available="No", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="Godrej Office Desk", category="Furniture", subcategory="Office Desks", vendor="Godrej Interio", purchase_date="2024-07-19", price=18999, serial_number="GOD-DESK-08", model_number="GI-DeskPro", invoice_number="INV-GI-2024-095", location="Finance Bay", assigned_user="Finance Team", description="Height-adjustable office desk", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", insurance_available="No", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="TV Unit Walnut Finish", category="Furniture", subcategory="TV Units", vendor="Pepperfry", purchase_date="2024-07-27", price=22499, serial_number="PF-TVU-443", model_number="Walnut-One", invoice_number="INV-PF-2024-109", location="Guest Lounge", assigned_user="Facilities", description="Lounge storage unit", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", insurance_available="No", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="Sony WH-1000XM5", category="Personal Gadgets", subcategory="Headphones", vendor="Sony Center", purchase_date="2024-08-03", price=29990, serial_number="SONY-XM5-77", model_number="WH1000XM5", invoice_number="INV-SONY-2024-121", location="Studio", assigned_user="Media Desk", description="Noise-cancelling headphones", warranty_provider="Sony", warranty_end_date="2025-08-02", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="Apple Watch Series 9", category="Personal Gadgets", subcategory="Smart Watches", vendor="Imagine", purchase_date="2024-08-11", price=42999, serial_number="AWS9-331", model_number="A2982", invoice_number="INV-AW-2024-136", location="Demo Kit", assigned_user="Product Team", description="Wearable for demo use", warranty_provider="Apple", warranty_end_date="2025-08-10", insurance_available="Yes", insurance_provider="HDFC Ergo", insurance_policy_number="POL-AW-1099", insurance_start_date="2024-08-11", insurance_expiry_date="2025-08-10", insurance_premium_amount=1999, insurance_coverage_notes="Damage protection", service_required="No", service_frequency="", service_reminder_enabled="No"),
    _sample_asset_row(product_name="Canon EOS R50", category="Personal Gadgets", subcategory="Cameras", vendor="PhotoWorld", purchase_date="2024-08-20", price=84999, serial_number="CAN-R50-22", model_number="EOS-R50", invoice_number="INV-CAN-2024-142", location="Studio", assigned_user="Content Team", description="Campaign camera body", warranty_provider="Canon India", warranty_end_date="2026-08-19", insurance_available="Yes", insurance_provider="Bajaj Allianz", insurance_policy_number="POL-CAM-7821", insurance_start_date="2024-08-20", insurance_expiry_date="2025-08-19", insurance_premium_amount=3600, insurance_coverage_notes="Lens and body coverage", service_frequency="quarterly"),
    _sample_asset_row(product_name="Hero Electric Scooter", category="Vehicles", subcategory="Electric Vehicles", vendor="Hero Electric", purchase_date="2024-09-02", price=112000, serial_number="HEV-2209", model_number="Optima CX", invoice_number="INV-HE-2024-151", location="Parking B2", assigned_user="Office Runner", description="Intra-campus transport", notes="Charged nightly", warranty_provider="Hero Electric", warranty_end_date="2027-09-01", insurance_available="Yes", insurance_provider="ICICI Lombard", insurance_policy_number="POL-EV-5510", insurance_start_date="2024-09-02", insurance_expiry_date="2025-09-01", insurance_premium_amount=6200, insurance_coverage_notes="Third-party + own damage", service_frequency="half_yearly"),
    _sample_asset_row(product_name="Maruti Ertiga", category="Vehicles", subcategory="Cars", vendor="Nexa", purchase_date="2024-09-14", price=1145000, serial_number="ERT-2024-01", model_number="ZXI+", invoice_number="INV-NEXA-2024-177", location="Company Fleet", assigned_user="Transport Team", description="Staff commute vehicle", warranty_provider="Maruti Suzuki", warranty_end_date="2026-09-13", insurance_available="Yes", insurance_provider="Tata AIG", insurance_policy_number="POL-CAR-2201", insurance_start_date="2024-09-14", insurance_expiry_date="2025-09-13", insurance_premium_amount=28750, insurance_coverage_notes="Comprehensive cover", service_frequency="quarterly"),
    _sample_asset_row(product_name="Firefox Commuter Cycle", category="Vehicles", subcategory="Bicycles", vendor="Decathlon", purchase_date="2024-09-26", price=18999, serial_number="FFC-1092", model_number="Urban 700C", invoice_number="INV-DCT-2024-188", location="Cycle Stand", assigned_user="Security Desk", description="Campus movement bicycle", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", insurance_available="No", service_frequency="custom", service_custom_interval_days=120),
    _sample_asset_row(product_name="Executive Visitor Sofa", category="Furniture", subcategory="Sofas", vendor="Home Centre", purchase_date="2024-10-04", price=38999, serial_number="SOFA-VC-09", model_number="Lounge-3S", invoice_number="INV-HC-2024-201", location="Reception Lounge", assigned_user="Front Office", description="3-seater visitor sofa", warranty_available="No", warranty_provider="", warranty_type="", warranty_start_date="", warranty_end_date="", warranty_notes="", warranty_reminder_30_days="No", warranty_reminder_7_days="No", warranty_reminder_on_expiry="No", insurance_available="No", service_required="No", service_frequency="", service_reminder_enabled="No"),
]


def _normalized_header_name(value: Any) -> str:
    text = str(value or "").lower().replace("_", " ")
    normalized = "".join(character if character.isalnum() or character == " " else " " for character in text)
    return " ".join(normalized.split())


def _header_aliases_for_field(field: dict[str, Any]) -> set[str]:
    key = str(field["key"])
    label = str(field["label"])
    aliases = {
        _normalized_header_name(key),
        _normalized_header_name(label),
    }
    if key == "price":
        aliases.add(_normalized_header_name("Price"))
    if key == "insurance_coverage_notes":
        aliases.add(_normalized_header_name("Coverage Notes"))
    return aliases


EXCEL_TEMPLATE_HEADER_ALIASES = {
    field["key"]: _header_aliases_for_field(field)
    for field in EXCEL_TEMPLATE_FIELDS
}
EXCEL_OPTIONAL_STATUS_HEADER_ALIASES = {
    _normalized_header_name("status"),
    _normalized_header_name("asset_status"),
    _normalized_header_name("asset status"),
}


def _template_row_values(row: dict[str, Any]) -> list[Any]:
    return [row.get(column, "") for column in EXCEL_TEMPLATE_COLUMNS]


def _excel_named_range_name(category_name: str, index: int) -> str:
    sanitized = "".join(character if character.isalnum() else "_" for character in category_name.strip())
    sanitized = sanitized.strip("_") or f"Category_{index}"
    if sanitized[0].isdigit():
        sanitized = f"Category_{sanitized}"
    return f"SubCategory_{index}_{sanitized}"[:255]


def _add_defined_name(workbook: Workbook, name: str, attr_text: str) -> None:
    defined_name = DefinedName(name=name, attr_text=attr_text)
    try:
        workbook.defined_names.add(defined_name)
    except AttributeError:
        workbook.defined_names.append(defined_name)


async def _get_template_category_map(db) -> dict[str, list[str]]:
    try:
        category_rows = await db["categories"].find({"is_active": {"$ne": False}}).sort("category", 1).to_list(length=1000)
        category_ids: list[str] = []
        category_names: dict[str, str] = {}
        for row in category_rows:
            category_id = str(row.get("_id", "")).strip()
            category_name = _text(row.get("name") or row.get("category"))
            if not category_id or not category_name:
                continue
            category_ids.append(category_id)
            category_names[category_id] = category_name

        if category_ids:
            subcategory_rows = await db["subcategories"].find(
                {"category_id": {"$in": category_ids}, "is_active": {"$ne": False}}
            ).to_list(length=5000)
            subcategory_map: dict[str, list[str]] = {category_id: [] for category_id in category_ids}
            for row in subcategory_rows:
                category_id = str(row.get("category_id", "")).strip()
                subcategory_name = _text(row.get("name"))
                if not category_id or not subcategory_name:
                    continue
                subcategory_map.setdefault(category_id, []).append(subcategory_name)

            category_map = {
                category_names[category_id]: sorted({*_dedupe_case_insensitive(subcategory_map.get(category_id, []))}, key=str.lower)
                for category_id in category_ids
                if category_id in category_names
            }
            if category_map:
                return category_map
    except Exception:
        pass

    return {
        category: list(subcategories)
        for category, subcategories in FINAL_CATEGORY_SUBCATEGORIES.items()
    }


def _dedupe_case_insensitive(values: list[str]) -> list[str]:
    unique: list[str] = []
    seen: set[str] = set()
    for raw in values:
        value = _text(raw)
        if not value:
            continue
        normalized = value.lower()
        if normalized in seen:
            continue
        seen.add(normalized)
        unique.append(value)
    return unique


def _apply_template_styles(sheet) -> None:
    for cell in sheet[1]:
        cell.font = EXCEL_TEMPLATE_HEADER_FONT
        cell.fill = EXCEL_TEMPLATE_HEADER_FILL

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for index, column_cells in enumerate(sheet.columns, start=1):
        values = [str(cell.value) for cell in column_cells if cell.value not in (None, "")]
        width = max((len(value) for value in values), default=12)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 14), 34)


def _apply_template_number_formats(sheet) -> None:
    for field_index, field in enumerate(EXCEL_TEMPLATE_FIELDS, start=1):
        column_letter = get_column_letter(field_index)
        if field["type"] == "date":
            for row_index in range(2, sheet.max_row + 1):
                sheet[f"{column_letter}{row_index}"].number_format = "yyyy-mm-dd"
        if field["type"] == "number":
            for row_index in range(2, sheet.max_row + 1):
                sheet[f"{column_letter}{row_index}"].number_format = "#,##0.00"


def _build_master_data_sheet(workbook: Workbook, category_map: dict[str, list[str]]) -> None:
    master_sheet = workbook.create_sheet(EXCEL_TEMPLATE_MASTER_SHEET_NAME)
    master_sheet.sheet_state = "hidden"

    categories = sorted(_dedupe_case_insensitive(list(category_map.keys())), key=str.lower)
    all_subcategories = _dedupe_case_insensitive(
        sorted(
            {subcategory for subcategories in category_map.values() for subcategory in subcategories},
            key=str.lower,
        )
    )

    master_sheet["A1"] = "Categories"
    for row_index, category in enumerate(categories, start=2):
        master_sheet[f"A{row_index}"] = category

    master_sheet["C1"] = "Category"
    master_sheet["D1"] = "NamedRange"
    for row_index, category in enumerate(categories, start=2):
        master_sheet[f"C{row_index}"] = category
        master_sheet[f"D{row_index}"] = _excel_named_range_name(category, row_index - 1)

    master_sheet["E1"] = "AllSubcategories"
    for row_index, subcategory in enumerate(all_subcategories, start=2):
        master_sheet[f"E{row_index}"] = subcategory

    quoted_sheet_name = quote_sheetname(EXCEL_TEMPLATE_MASTER_SHEET_NAME)
    if categories:
        _add_defined_name(workbook, "CategoryOptions", f"{quoted_sheet_name}!$A$2:$A${len(categories) + 1}")
    if all_subcategories:
        _add_defined_name(workbook, "AllSubcategories", f"{quoted_sheet_name}!$E$2:$E${len(all_subcategories) + 1}")

    for offset, category in enumerate(categories, start=6):
        column_letter = get_column_letter(offset)
        subcategories = _dedupe_case_insensitive(category_map.get(category, []))
        master_sheet[f"{column_letter}1"] = category
        for row_index, subcategory in enumerate(subcategories, start=2):
            master_sheet[f"{column_letter}{row_index}"] = subcategory

        if subcategories:
            named_range = _excel_named_range_name(category, offset - 5)
            _add_defined_name(
                workbook,
                named_range,
                f"{quoted_sheet_name}!${column_letter}$2:${column_letter}${len(subcategories) + 1}",
            )


def _apply_template_validations(sheet, workbook: Workbook, category_map: dict[str, list[str]]) -> None:
    category_index = EXCEL_TEMPLATE_COLUMNS.index("category") + 1
    subcategory_index = EXCEL_TEMPLATE_COLUMNS.index("subcategory") + 1
    category_column = get_column_letter(category_index)

    data_end_row = max(EXCEL_TEMPLATE_ROW_LIMIT, sheet.max_row)

    list_validations = {
        "warranty_available": '"Yes,No"',
        "warranty_type": '"manufacturer,extended"',
        "warranty_reminder_30_days": '"Yes,No"',
        "warranty_reminder_7_days": '"Yes,No"',
        "warranty_reminder_on_expiry": '"Yes,No"',
        "insurance_available": '"Yes,No"',
        "insurance_reminder_45_days": '"Yes,No"',
        "insurance_reminder_15_days": '"Yes,No"',
        "service_required": '"Yes,No"',
        "service_frequency": '"monthly,quarterly,half_yearly,yearly,custom"',
        "service_reminder_enabled": '"Yes,No"',
    }

    for key, formula in list_validations.items():
        field_index = EXCEL_TEMPLATE_COLUMNS.index(key) + 1
        column_letter = get_column_letter(field_index)
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{data_end_row}")

    category_validation = DataValidation(type="list", formula1="=CategoryOptions", allow_blank=False)
    sheet.add_data_validation(category_validation)
    category_validation.add(f"{category_column}2:{category_column}{data_end_row}")

    mapping_end_row = len(category_map) + 1
    subcategory_validation_formula = (
        f'=INDIRECT(IFERROR(VLOOKUP(${category_column}2,{quote_sheetname(EXCEL_TEMPLATE_MASTER_SHEET_NAME)}!$C$2:$D${mapping_end_row},2,FALSE),"AllSubcategories"))'
    )
    subcategory_validation = DataValidation(type="list", formula1=subcategory_validation_formula, allow_blank=False)
    sheet.add_data_validation(subcategory_validation)
    subcategory_column = get_column_letter(subcategory_index)
    subcategory_validation.add(f"{subcategory_column}2:{subcategory_column}{data_end_row}")

    for field in EXCEL_TEMPLATE_FIELDS:
        field_index = EXCEL_TEMPLATE_COLUMNS.index(field["key"]) + 1
        column_letter = get_column_letter(field_index)
        if field["type"] == "date":
            validation = DataValidation(type="date", operator="between", formula1="DATE(1900,1,1)", formula2="DATE(9999,12,31)", allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}{data_end_row}")
        if field["type"] == "number":
            validation = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}{data_end_row}")


def _normalize_excel_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized_row = dict(row)

    category = _text(normalized_row.get("category"))
    custom_category = _text(normalized_row.get("custom_category"))
    if category.lower() == "other" and custom_category:
        normalized_row["category"] = custom_category

    subcategory = _text(normalized_row.get("subcategory"))
    custom_subcategory = _text(normalized_row.get("custom_subcategory"))
    if subcategory.lower() == "other" and custom_subcategory:
        normalized_row["subcategory"] = custom_subcategory

    return normalized_row


def _to_asset(item: dict[str, Any]) -> dict[str, Any]:
    source = _normalize_asset_source(item.get("source"))
    return {
        "id": str(item.get("_id", "")),
        "name": item.get("name", ""),
        "asset_name": item.get("asset_name") or item.get("name"),
        "status": item.get("status"),
        "vendor": item.get("vendor"),
        "purchase_date": item.get("purchase_date"),
        "price": item.get("price"),
        "source": source,
        "user_id": item.get("user_id", ""),
        "brand": item.get("brand"),
        "category": item.get("category"),
        "subcategory": item.get("subcategory"),
        "serial_number": item.get("serial_number"),
        "model_number": item.get("model_number"),
        "invoice_number": item.get("invoice_number"),
        "description": item.get("description"),
        "notes": item.get("notes"),
        "location": item.get("location"),
        "assigned_user": item.get("assigned_user"),
        "warranty": item.get("warranty"),
        "insurance": item.get("insurance"),
        "service": item.get("service"),
        "source_email_id": item.get("source_email_id"),
        "source_email_sender": item.get("source_email_sender"),
        "source_email_subject": item.get("source_email_subject"),
        "invoice_attachment_path": item.get("invoice_attachment_path"),
        "auto_reminders_created": item.get("auto_reminders_created", 0),
        "created_at": item.get("created_at"),
        "updated_at": item.get("updated_at"),
    }


def _to_document(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "document_id": str(item.get("_id", "")),
        "file_name": item.get("file_name", "document"),
        "document_type": item.get("document_type", "supporting"),
        "file_url": f"/api/assets/{item.get('asset_id')}/documents/{item.get('_id')}/file",
        "uploaded_at": item.get("uploaded_at", datetime.now(timezone.utc).isoformat()),
    }


def _text(value: Any) -> str:
    return str(value or "").strip()


def _normalized_lower(value: Any) -> str:
    return _text(value).lower()


def _normalize_asset_source(value: Any) -> str:
    normalized = _normalized_lower(value)
    return ASSET_SOURCE_ALIASES.get(normalized, ASSET_SOURCE_MANUAL)


def _is_truthy(value: Any) -> bool:
    normalized = _normalized_lower(value)
    return normalized in {"true", "1", "yes", "y", "on"}


def _to_number(value: Any) -> float | None:
    if value is None:
        return None
    text = _text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_int(value: Any) -> int | None:
    number = _to_number(value)
    if number is None:
        return None
    return int(number)


def _to_iso_date_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    text = _text(value)
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text)
        return parsed.date().isoformat()
    except ValueError:
        return text


def _is_valid_date_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, (datetime, date)):
        return True

    text = _text(value)
    if not text:
        return True

    normalized = text.replace("Z", "+00:00")
    try:
        datetime.fromisoformat(normalized)
        return True
    except ValueError:
        pass

    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue

    return False


def _parse_date(value: Any) -> date | None:
    """Parse any value to a date object."""
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    
    text = _text(value)
    if not text:
        return None
    
    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
        return parsed.date()
    except ValueError:
        pass
    
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    
    return None


def _extract_date_field(data: dict[str, Any] | None, *keys: str) -> date | None:
    """Extract a date from nested dict, trying multiple key names."""
    if not isinstance(data, dict):
        return None
    for key in keys:
        value = data.get(key)
        parsed = _parse_date(value)
        if parsed:
            return parsed
    return None


def _extract_boolean_field(data: dict[str, Any] | None, *keys: str) -> bool:
    """Extract a boolean from nested dict, trying multiple key names."""
    if not isinstance(data, dict):
        return False
    for key in keys:
        value = data.get(key)
        if _is_truthy(value):
            return True
    return False


def _compute_asset_status(asset_data: dict[str, Any]) -> str:
    """
    Compute asset status from lifecycle data and is_inactive flag.
    
    Rules (in order):
    1. If is_inactive=true → "Inactive"
    2. If warranty/insurance/service expired → "Expired"
    3. If warranty/insurance/service expiring within 30 days → "Expiring Soon"
    4. If warranty/insurance/service valid → "In Warranty"
    5. Otherwise → "Active"
    """
    # Rule 1: Check if explicitly marked inactive
    if asset_data.get("is_inactive") is True:
        return "Inactive"
    
    # Extract lifecycle info
    warranty = asset_data.get("warranty")
    insurance = asset_data.get("insurance")
    service = asset_data.get("service")
    
    # Collect all relevant dates
    dates_to_check: list[date] = []
    
    # Extract warranty end date
    if _extract_boolean_field(warranty, "available"):
        warranty_end = _extract_date_field(warranty, "end_date", "expiry_date")
        if warranty_end:
            dates_to_check.append(warranty_end)
    
    # Extract insurance expiry date
    if _extract_boolean_field(insurance, "available"):
        insurance_end = _extract_date_field(insurance, "expiry_date", "end_date")
        if insurance_end:
            dates_to_check.append(insurance_end)
    
    # Extract service next due date
    if _extract_boolean_field(service, "required"):
        service_due = _extract_date_field(service, "next_service_date", "next_due_date")
        if service_due:
            dates_to_check.append(service_due)
    
    # If no dates, asset is active
    if not dates_to_check:
        return "Active"
    
    # Find nearest date
    today = date.today()
    nearest = min(dates_to_check)
    
    # Rule 2: Check if expired
    if nearest < today:
        return "Expired"
    
    # Rule 3: Check if expiring soon (within 30 days)
    days_until = (nearest - today).days
    if days_until <= 30:
        return "Expiring Soon"
    
    # Rule 4: Valid warranty/insurance/service
    return "In Warranty"


async def _get_upload_validation_category_map(db) -> tuple[dict[str, str], dict[str, set[str]]]:
    category_lookup: dict[str, str] = {}
    subcategory_lookup: dict[str, set[str]] = {}

    try:
        category_rows = await db["categories"].find({"is_active": {"$ne": False}}).to_list(length=1000)
        category_id_lookup: dict[str, str] = {}
        category_id_to_key: dict[str, str] = {}
        for row in category_rows:
            category_id = str(row.get("_id", "")).strip()
            category_name = _text(row.get("name") or row.get("category"))
            if not category_id or not category_name:
                continue
            category_key = category_name.lower()
            category_lookup[category_key] = category_name
            category_id_lookup[category_id] = category_name
            category_id_to_key[category_id] = category_key
            subcategory_lookup.setdefault(category_key, set())

        if category_id_lookup:
            subcategory_rows = await db["subcategories"].find(
                {
                    "category_id": {"$in": list(category_id_lookup.keys())},
                    "is_active": {"$ne": False},
                }
            ).to_list(length=5000)
            for row in subcategory_rows:
                category_id = str(row.get("category_id", "")).strip()
                subcategory_name = _text(row.get("name"))
                category_key = category_id_to_key.get(category_id)
                if not category_key or not subcategory_name:
                    continue
                subcategory_lookup.setdefault(category_key, set()).add(subcategory_name.lower())

        if category_lookup:
            return category_lookup, subcategory_lookup
    except Exception:
        pass

    for category, subcategories in FINAL_CATEGORY_SUBCATEGORIES.items():
        category_key = _text(category).lower()
        if not category_key:
            continue
        category_lookup[category_key] = _text(category)
        subcategory_lookup[category_key] = {
            _text(subcategory).lower()
            for subcategory in subcategories
            if _text(subcategory)
        }

    return category_lookup, subcategory_lookup


def _validate_excel_row(
    row: dict[str, Any],
    category_lookup: dict[str, str],
    subcategory_lookup: dict[str, set[str]],
) -> list[str]:
    errors: list[str] = []

    product_name = _text(row.get("product_name"))
    if not product_name:
        errors.append("Asset Name is required")

    category = _text(row.get("category"))
    subcategory = _text(row.get("subcategory"))
    canonical_category_key = category.lower() if category else ""

    if not category:
        errors.append("Category is required")
    elif canonical_category_key not in category_lookup:
        errors.append("Invalid category")

    if not subcategory:
        errors.append("SubCategory is required")
    elif canonical_category_key in subcategory_lookup:
        if subcategory.lower() not in subcategory_lookup.get(canonical_category_key, set()):
            errors.append("Invalid subcategory for selected category")

    purchase_date_value = row.get("purchase_date")
    if _text(purchase_date_value) and not _is_valid_date_value(purchase_date_value):
        errors.append("Purchase Date must be a valid date")

    purchase_price_value = row.get("price")
    if _text(purchase_price_value) and _to_number(purchase_price_value) is None:
        errors.append("Purchase Price must be numeric")

    warranty_start_date = row.get("warranty_start_date")
    if _text(warranty_start_date) and not _is_valid_date_value(warranty_start_date):
        errors.append("Warranty Start Date must be a valid date")

    warranty_end_date = row.get("warranty_end_date")
    if _text(warranty_end_date) and not _is_valid_date_value(warranty_end_date):
        errors.append("Warranty End Date must be a valid date")

    return errors


def _to_excel_suggestion(
    row: dict[str, Any],
    row_number: int,
    duplicate_asset_id: str | None,
) -> dict[str, Any]:
    warranty_available = _is_truthy(row.get("warranty_available"))
    insurance_available = _is_truthy(row.get("insurance_available"))
    service_required = _is_truthy(row.get("service_required"))

    warranty = {
        "available": warranty_available,
        "provider": _text(row.get("warranty_provider")) or None,
        "type": _text(row.get("warranty_type")) or "manufacturer",
        "start_date": _to_iso_date_text(row.get("warranty_start_date")),
        "end_date": _to_iso_date_text(row.get("warranty_end_date")),
        "notes": _text(row.get("warranty_notes")) or None,
        "reminders": {
            "thirty_days_before": _is_truthy(row.get("warranty_reminder_30_days")),
            "seven_days_before": _is_truthy(row.get("warranty_reminder_7_days")),
            "on_expiry": _is_truthy(row.get("warranty_reminder_on_expiry")),
        },
    }

    insurance = {
        "available": insurance_available,
        "provider": _text(row.get("insurance_provider")) or None,
        "policy_number": _text(row.get("insurance_policy_number")) or None,
        "start_date": _to_iso_date_text(row.get("insurance_start_date")),
        "expiry_date": _to_iso_date_text(row.get("insurance_expiry_date")),
        "premium_amount": _to_number(row.get("insurance_premium_amount")),
        "coverage_notes": _text(row.get("insurance_coverage_notes")) or None,
        "reminders": {
            "forty_five_days_before": _is_truthy(row.get("insurance_reminder_45_days")),
            "fifteen_days_before": _is_truthy(row.get("insurance_reminder_15_days")),
        },
    }

    service = {
        "required": service_required,
        "frequency": _text(row.get("service_frequency")) or "monthly",
        "custom_interval_days": _to_int(row.get("service_custom_interval_days")),
        "reminder_enabled": _is_truthy(row.get("service_reminder_enabled")),
    }

    now = datetime.now(timezone.utc)
    status = "duplicate" if duplicate_asset_id else "new"
    return {
        "id": f"excel-row-{row_number}",
        "product_name": _text(row.get("product_name")) or f"Asset Row {row_number}",
        "brand": _text(row.get("brand")) or None,
        "vendor": _text(row.get("vendor")) or None,
        "price": _to_number(row.get("price")),
        "purchase_date": _to_iso_date_text(row.get("purchase_date")),
        "sender": "Excel Upload",
        "subject": "Bulk Asset Import",
        "email_date": now.isoformat(),
        "quantity": 1,
        "source": "excel",
        "status": status,
        "warranty": _text(row.get("warranty_notes")) or None,
        "email_message_id": f"excel-upload-{row_number}",
        "attachment_filename": None,
        "attachment_mime_type": None,
        "action_options": ["add", "skip"],
        "already_added": bool(duplicate_asset_id),
        "created_at": now.isoformat(),
        "category": _text(row.get("category")) or "Other",
        "subcategory": _text(row.get("subcategory")) or "Custom Asset",
        "serial_number": _text(row.get("serial_number")) or None,
        "model_number": _text(row.get("model_number")) or None,
        "invoice_number": _text(row.get("invoice_number")) or None,
        "description": _text(row.get("description")) or None,
        "notes": _text(row.get("notes")) or None,
        "location": _text(row.get("location")) or None,
        "assigned_user": _text(row.get("assigned_user")) or None,
        "warranty_details": warranty if warranty_available else None,
        "insurance_details": insurance if insurance_available else None,
        "service_details": service if service_required else None,
        "asset_id": duplicate_asset_id,
    }


def _lifecycle_payload(payload: dict[str, Any]) -> dict[str, Any] | None:
    lifecycle = payload.get("lifecycle_info")
    if isinstance(lifecycle, dict):
        return lifecycle
    return None


def _enrich_service_lifecycle(service: dict[str, Any] | None) -> dict[str, Any] | None:
    """Compute and persist next_service_date into the service dict before saving."""
    if not isinstance(service, dict) or not service.get("required"):
        return service
    interval_days = int(service.get("custom_interval_days") or 0)
    if interval_days <= 0:
        frequency = str(service.get("frequency") or "monthly")
        mapping = {"monthly": 30, "quarterly": 90, "half_yearly": 180, "yearly": 365}
        interval_days = mapping.get(frequency, 30)
    next_service_date = (datetime.now(timezone.utc) + timedelta(days=interval_days)).date().isoformat()
    return {**service, "next_service_date": next_service_date}


async def _create_reminders_for_lifecycle(
    asset_id: str,
    asset_name: str,
    lifecycle_info: dict[str, Any] | None,
    user_id: str,
    db,
) -> int:
    if not lifecycle_info:
        return 0

    reminders: list[dict[str, Any]] = []
    now_iso = datetime.now(timezone.utc).isoformat()

    warranty = lifecycle_info.get("warranty")
    if isinstance(warranty, dict) and warranty.get("available") and warranty.get("end_date"):
        reminders.append(
            {
                "user_id": user_id,
                "title": f"Warranty expiry for {asset_name}",
                "asset_id": asset_id,
                "asset_name": asset_name,
                "reminder_date": str(warranty["end_date"]),
                "reminder_type": "warranty",
                "status": "active",
                "notes": warranty.get("notes"),
                "created_at": now_iso,
                "updated_at": now_iso,
            }
        )

    insurance = lifecycle_info.get("insurance")
    if isinstance(insurance, dict) and insurance.get("available") and insurance.get("expiry_date"):
        reminders.append(
            {
                "user_id": user_id,
                "title": f"Insurance renewal for {asset_name}",
                "asset_id": asset_id,
                "asset_name": asset_name,
                "reminder_date": str(insurance["expiry_date"]),
                "reminder_type": "custom",
                "status": "active",
                "notes": insurance.get("coverage_notes") or insurance.get("notes"),
                "created_at": now_iso,
                "updated_at": now_iso,
            }
        )

    service = lifecycle_info.get("service")
    if isinstance(service, dict) and service.get("required"):
        interval_days = int(service.get("custom_interval_days") or 0)
        if interval_days <= 0:
            frequency = str(service.get("frequency") or "monthly")
            mapping = {"monthly": 30, "quarterly": 90, "half_yearly": 180, "yearly": 365}
            interval_days = mapping.get(frequency, 30)

        reminder_date = (datetime.now(timezone.utc) + timedelta(days=interval_days)).date().isoformat()
        reminders.append(
            {
                "user_id": user_id,
                "title": f"Service reminder for {asset_name}",
                "asset_id": asset_id,
                "asset_name": asset_name,
                "reminder_date": reminder_date,
                "reminder_type": "service",
                "status": "active",
                "notes": "Auto-generated from asset service settings",
                "created_at": now_iso,
                "updated_at": now_iso,
            }
        )

    if reminders:
        await db["reminders"].insert_many(reminders)

    return len(reminders)


async def _sync_reminders_on_asset_update(
    asset_id: str,
    asset_name: str,
    old_lifecycle: dict[str, Any] | None,
    new_lifecycle: dict[str, Any] | None,
    user_id: str,
    db,
) -> None:
    """
    Sync reminders when asset lifecycle details are updated.
    Deletes old reminders for changed fields and creates new ones.
    """
    if not new_lifecycle:
        return

    # Determine which lifecycle fields changed
    old_warranty = (old_lifecycle.get("warranty") if old_lifecycle else None) or {}
    new_warranty = new_lifecycle.get("warranty") or {}
    warranty_changed = old_warranty != new_warranty

    old_insurance = (old_lifecycle.get("insurance") if old_lifecycle else None) or {}
    new_insurance = new_lifecycle.get("insurance") or {}
    insurance_changed = old_insurance != new_insurance

    old_service = (old_lifecycle.get("service") if old_lifecycle else None) or {}
    new_service = new_lifecycle.get("service") or {}
    service_changed = old_service != new_service

    # Delete old reminders for changed fields
    if warranty_changed:
        await db["reminders"].delete_many({"asset_id": asset_id, "user_id": user_id, "reminder_type": "warranty"})

    if insurance_changed:
        await db["reminders"].delete_many({"asset_id": asset_id, "user_id": user_id, "reminder_type": "custom", "title": {"$regex": f"Insurance renewal for {asset_name}"}})

    if service_changed:
        await db["reminders"].delete_many({"asset_id": asset_id, "user_id": user_id, "reminder_type": "service"})

    # Create new reminders based on updated lifecycle info
    await _create_reminders_for_lifecycle(asset_id, asset_name, new_lifecycle, user_id, db)


async def _find_duplicate_asset(payload: dict[str, Any], user_id: str, db) -> dict[str, Any] | None:
    invoice_number = _text(payload.get("invoice_number"))
    if invoice_number:
        duplicate = await db["assets"].find_one(
            {
                "user_id": user_id,
                "invoice_number": {"$regex": f"^{invoice_number}$", "$options": "i"},
            }
        )
        if duplicate:
            return duplicate

    name = _text(payload.get("name") or payload.get("product_name"))
    vendor = _text(payload.get("vendor"))
    purchase_date = _text(payload.get("purchase_date"))
    source_email_id = _text(payload.get("source_email_id"))

    if source_email_id:
        duplicate = await db["assets"].find_one({"user_id": user_id, "source_email_id": source_email_id})
        if duplicate:
            return duplicate

    if name and vendor and purchase_date:
        duplicate = await db["assets"].find_one(
            {
                "user_id": user_id,
                "name_normalized": _normalized_lower(name),
                "vendor_normalized": _normalized_lower(vendor),
                "purchase_date": purchase_date,
            }
        )
        if duplicate:
            return duplicate

    return None


@router.get("/excel/template")
async def download_excel_template(current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)):
    _ = current_user
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = EXCEL_TEMPLATE_SHEET_NAME
    sheet.append(EXCEL_TEMPLATE_HEADERS)

    for row in EXCEL_TEMPLATE_SAMPLE_ROWS:
        sheet.append(_template_row_values(row))

    category_map = await _get_template_category_map(db)
    _build_master_data_sheet(workbook, category_map)
    _apply_template_styles(sheet)
    _apply_template_number_formats(sheet)
    _apply_template_validations(sheet, workbook, category_map)

    payload = BytesIO()
    workbook.save(payload)
    payload.seek(0)

    return StreamingResponse(
        BytesIO(payload.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="asset_upload_template.xlsx"'},
    )


@router.post("/excel/upload")
async def upload_excel_file(
    file: UploadFile = File(...),
    current_user: dict[str, str] = Depends(get_current_user),
    db=Depends(get_db),
) -> dict[str, Any]:
    file_name = (file.filename or "").lower()
    if not file_name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file format or empty file")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Invalid file format or empty file")

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as error:
        raise HTTPException(status_code=400, detail="Invalid file format or empty file") from error

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Invalid file format or empty file")

    normalized_headers = [_normalized_header_name(header) for header in (rows[0] or [])]
    header_index_by_alias: dict[str, int] = {}
    for index, header in enumerate(normalized_headers):
        if header and header not in header_index_by_alias:
            header_index_by_alias[header] = index

    missing_headers = [
        EXCEL_TEMPLATE_HEADER_BY_KEY[column]
        for column in EXCEL_TEMPLATE_COLUMNS
        if not any(alias in header_index_by_alias for alias in EXCEL_TEMPLATE_HEADER_ALIASES[column])
    ]
    if missing_headers:
        raise HTTPException(
            status_code=400,
            detail="Invalid template. Please use the correct format",
        )

    header_index = {
        column: next(
            header_index_by_alias[alias]
            for alias in EXCEL_TEMPLATE_HEADER_ALIASES[column]
            if alias in header_index_by_alias
        )
        for column in EXCEL_TEMPLATE_COLUMNS
    }
    optional_status_index = next(
        (header_index_by_alias[alias] for alias in EXCEL_OPTIONAL_STATUS_HEADER_ALIASES if alias in header_index_by_alias),
        None,
    )
    category_lookup, subcategory_lookup = await _get_upload_validation_category_map(db)

    suggestions: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []
    row_results: list[dict[str, Any]] = []
    valid_rows = 0
    invalid_rows = 0

    for index, excel_row in enumerate(rows[1:], start=2):
        row_map: dict[str, Any] = {}
        for column in EXCEL_TEMPLATE_COLUMNS:
            position = header_index[column]
            row_map[column] = excel_row[position] if position < len(excel_row) else None

        row_map = _normalize_excel_row(row_map)

        if not any(_text(value) for value in row_map.values()):
            skipped_rows.append({"row_number": index, "reason": "Empty row"})
            continue

        validation_errors = _validate_excel_row(row_map, category_lookup, subcategory_lookup)
        is_valid = len(validation_errors) == 0

        duplicate_asset_id: str | None = None
        if is_valid:
            duplicate = await _find_duplicate_asset(
                {
                    "name": row_map.get("product_name"),
                    "vendor": row_map.get("vendor"),
                    "purchase_date": _to_iso_date_text(row_map.get("purchase_date")),
                    "invoice_number": row_map.get("invoice_number"),
                },
                current_user["id"],
                db,
            )
            duplicate_asset_id = str(duplicate.get("_id")) if duplicate else None

        suggestion = _to_excel_suggestion(row_map, index, duplicate_asset_id)
        suggestion["row_number"] = index - 1
        suggestion["validation_status"] = "valid" if is_valid else "invalid"
        suggestion["validation_errors"] = validation_errors
        if not is_valid:
            suggestion["status"] = "invalid"
            suggestion["already_added"] = False
            suggestion["asset_id"] = None

        suggestions.append(suggestion)

        row_results.append(
            {
                "row": index - 1,
                "status": "valid" if is_valid else "invalid",
                "errors": validation_errors,
            }
        )
        if is_valid:
            valid_rows += 1
        else:
            invalid_rows += 1

    return {
        "file_name": file.filename,
        "total": len(row_results),
        "valid": valid_rows,
        "invalid": invalid_rows,
        "data": row_results,
        "total_rows": max(len(rows) - 1, 0),
        "parsed_rows": valid_rows,
        "skipped_rows": skipped_rows,
        "suggestions": suggestions,
    }


@router.get("")
async def list_assets(current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)) -> list[dict[str, Any]]:
    items = await db["assets"].find({"user_id": current_user["id"]}).sort("created_at", -1).to_list(length=2000)
    return [_to_asset(item) for item in items]


@router.post("")
async def create_asset(payload: dict[str, Any], current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)) -> dict[str, Any]:
    category = _text(payload.get("category"))
    subcategory = _text(payload.get("subcategory"))
    if not category:
        raise HTTPException(status_code=400, detail="Category is required")
    if not subcategory:
        raise HTTPException(status_code=400, detail="SubCategory is required")

    duplicate = await _find_duplicate_asset(payload, current_user["id"], db)
    if duplicate:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Duplicate asset detected",
                "existing_asset_id": str(duplicate.get("_id", "")),
                "name": duplicate.get("name"),
                "invoice_number": duplicate.get("invoice_number"),
            },
        )

    now_iso = datetime.now(timezone.utc).isoformat()
    name = _text(payload.get("name") or payload.get("product_name")) or "Unnamed Asset"
    source = _normalize_asset_source(payload.get("source"))
    suggestion_id = payload.get("suggestion_id")
    suggestion_object_id: ObjectId | None = None
    suggestion_item: dict[str, Any] | None = None

    if suggestion_id:
        try:
            suggestion_object_id = ObjectId(str(suggestion_id))
            suggestion_item = await db["asset_suggestions"].find_one(
                {"_id": suggestion_object_id, "user_id": current_user["id"]}
            )
        except Exception:
            suggestion_object_id = None
            suggestion_item = None

    suggestion_attachment_path = _text(payload.get("invoice_attachment_path"))
    if not suggestion_attachment_path and suggestion_item:
        suggestion_attachment_path = _text(
            suggestion_item.get("invoice_attachment_path") or suggestion_item.get("attachment_path")
        )

    suggestion_file_name = None
    if suggestion_item:
        suggestion_file_name = _text(
            suggestion_item.get("attachment_filename")
            or suggestion_item.get("file_name")
            or suggestion_item.get("invoice_filename")
        )

    lifecycle_info = _lifecycle_payload(payload)
    enriched_service = _enrich_service_lifecycle(lifecycle_info.get("service")) if lifecycle_info else None
    
    # Prepare document with enriched lifecycle data for status computation
    temp_asset_for_status = {
        "is_inactive": payload.get("is_inactive", False),
        "warranty": lifecycle_info.get("warranty") if lifecycle_info else None,
        "insurance": lifecycle_info.get("insurance") if lifecycle_info else None,
        "service": enriched_service,
    }
    computed_status = _compute_asset_status(temp_asset_for_status)
    
    document = {
        "user_id": current_user["id"],
        "name": name,
        "name_normalized": _normalized_lower(name),
        "asset_name": name,
        "is_inactive": payload.get("is_inactive", False),
        "status": computed_status,
        "brand": payload.get("brand"),
        "category": category,
        "subcategory": subcategory,
        "vendor": payload.get("vendor"),
        "vendor_normalized": _normalized_lower(payload.get("vendor")),
        "purchase_date": payload.get("purchase_date"),
        "price": payload.get("price"),
        "serial_number": payload.get("serial_number"),
        "model_number": payload.get("model_number"),
        "invoice_number": _text(payload.get("invoice_number")) or None,
        "description": payload.get("description"),
        "notes": payload.get("notes"),
        "location": payload.get("location"),
        "assigned_user": payload.get("assigned_user"),
        "warranty": lifecycle_info.get("warranty") if lifecycle_info else None,
        "insurance": lifecycle_info.get("insurance") if lifecycle_info else None,
        "service": enriched_service,
        "source": source,
        "source_email_id": payload.get("source_email_id"),
        "source_email_sender": payload.get("source_email_sender"),
        "source_email_subject": payload.get("source_email_subject"),
        "invoice_attachment_path": suggestion_attachment_path,
        "auto_reminders_created": 0,
        "created_at": now_iso,
        "updated_at": now_iso,
    }

    result = await db["assets"].insert_one(document)
    asset_id = str(result.inserted_id)

    if suggestion_id:
        try:
            object_id = suggestion_object_id or ObjectId(str(suggestion_id))
            await db["asset_suggestions"].update_one(
                {"_id": object_id, "user_id": current_user["id"]},
                {"$set": {"already_added": True, "status": "confirmed", "asset_id": asset_id, "updated_at": now_iso}},
            )
        except Exception:
            pass

    if suggestion_id and suggestion_attachment_path:
        existing_invoice_doc = await db["asset_documents"].find_one(
            {
                "asset_id": asset_id,
                "user_id": current_user["id"],
                "file_path": suggestion_attachment_path,
            }
        )
        if not existing_invoice_doc:
            invoice_file_name = suggestion_file_name or suggestion_attachment_path.split("/")[-1] or "Invoice"
            await db["asset_documents"].insert_one(
                {
                    "asset_id": asset_id,
                    "user_id": current_user["id"],
                    "file_name": invoice_file_name,
                    "file_path": suggestion_attachment_path,
                    "document_type": "invoice",
                    "uploaded_at": now_iso,
                }
            )

    reminder_count = await _create_reminders_for_lifecycle(asset_id, name, lifecycle_info, current_user["id"], db)
    if reminder_count:
        await db["assets"].update_one({"_id": ObjectId(asset_id)}, {"$set": {"auto_reminders_created": reminder_count}})

    created = await db["assets"].find_one({"_id": ObjectId(asset_id)})
    return _to_asset(created or document)


@router.get("/{asset_id}")
async def get_asset(asset_id: str, current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)) -> dict[str, Any]:
    try:
        object_id = ObjectId(asset_id)
    except Exception as error:  # pragma: no cover
        raise HTTPException(status_code=400, detail="Invalid asset id") from error

    item = await db["assets"].find_one({"_id": object_id, "user_id": current_user["id"]})
    if not item:
        raise HTTPException(status_code=404, detail="Asset not found")
    return _to_asset(item)


@router.put("/{asset_id}")
async def update_asset(asset_id: str, payload: dict[str, Any], current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)) -> dict[str, Any]:
    try:
        object_id = ObjectId(asset_id)
    except Exception as error:  # pragma: no cover
        raise HTTPException(status_code=400, detail="Invalid asset id") from error

    # Fetch old asset to detect lifecycle changes
    old_asset = await db["assets"].find_one({"_id": object_id, "user_id": current_user["id"]})
    if not old_asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    lifecycle_info = _lifecycle_payload(payload)
    allowed = {
        "name",
        "asset_name",
        "is_inactive",
        "status",
        "brand",
        "vendor",
        "purchase_date",
        "price",
        "category",
        "subcategory",
        "serial_number",
        "model_number",
        "invoice_number",
        "description",
        "notes",
        "location",
        "assigned_user",
        "warranty",
        "insurance",
        "service",
        "source",
    }
    update_data = {key: value for key, value in payload.items() if key in allowed}

    if "name" in update_data:
        update_data["name_normalized"] = _normalized_lower(update_data.get("name"))
    if "vendor" in update_data:
        update_data["vendor_normalized"] = _normalized_lower(update_data.get("vendor"))

    if lifecycle_info:
        update_data["warranty"] = lifecycle_info.get("warranty")
        update_data["insurance"] = lifecycle_info.get("insurance")
        update_data["service"] = _enrich_service_lifecycle(lifecycle_info.get("service"))

    category = update_data.get("category")
    if category is not None and not _text(category):
        raise HTTPException(status_code=400, detail="Category cannot be empty")

    subcategory = update_data.get("subcategory")
    if subcategory is not None and not _text(subcategory):
        raise HTTPException(status_code=400, detail="SubCategory cannot be empty")


    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    # Determine status to store.
    # "Inactive", "Lost", "Damaged" are explicit user-controlled statuses that should be
    # stored as-is. Lifecycle statuses ("Active", "In Warranty", etc.) are always
    # recomputed from lifecycle dates so they stay accurate.
    _explicit_statuses = {"Inactive", "Lost", "Damaged"}
    explicit_status = _text(payload.get("status"))
    if explicit_status and explicit_status in _explicit_statuses:
        # User explicitly chose a non-lifecycle status — preserve it and sync is_inactive.
        update_data["status"] = explicit_status
        update_data["is_inactive"] = (explicit_status == "Inactive")
    else:
        # Lifecycle status (or no status sent): recompute from merged data.
        # If the user explicitly picked a lifecycle status (e.g. "Active"), clear
        # is_inactive first so a previously-inactive asset is not stuck as Inactive.
        if explicit_status:
            update_data["is_inactive"] = False
        merged_asset_data = {**old_asset, **update_data}
        update_data["status"] = _compute_asset_status(merged_asset_data)
        update_data["is_inactive"] = update_data["status"] == "Inactive"
    await db["assets"].update_one({"_id": object_id, "user_id": current_user["id"]}, {"$set": update_data})
    item = await db["assets"].find_one({"_id": object_id, "user_id": current_user["id"]})
    if not item:
        raise HTTPException(status_code=404, detail="Asset not found")

    # Sync reminders if lifecycle info was updated
    if lifecycle_info:
        asset_name = str(item.get("name") or "Asset")
        old_lifecycle = (
            {
                "warranty": old_asset.get("warranty"),
                "insurance": old_asset.get("insurance"),
                "service": old_asset.get("service"),
            }
            if old_asset
            else None
        )
        await _sync_reminders_on_asset_update(asset_id, asset_name, old_lifecycle, lifecycle_info, current_user["id"], db)

    return _to_asset(item)


@router.delete("/{asset_id}")
async def delete_asset(asset_id: str, current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)) -> dict[str, str]:
    try:
        object_id = ObjectId(asset_id)
    except Exception as error:  # pragma: no cover
        raise HTTPException(status_code=400, detail="Invalid asset id") from error

    docs = await db["asset_documents"].find({"asset_id": asset_id, "user_id": current_user["id"]}).to_list(length=300)
    for doc in docs:
        path = Path(str(doc.get("file_path", "")))
        if path.exists():
            path.unlink(missing_ok=True)

    await db["asset_documents"].delete_many({"asset_id": asset_id, "user_id": current_user["id"]})
    await db["reminders"].delete_many({"asset_id": asset_id, "user_id": current_user["id"]})
    await db["assets"].delete_one({"_id": object_id, "user_id": current_user["id"]})
    return {"status": "deleted", "asset_id": asset_id}


@router.get("/{asset_id}/invoice")
async def get_asset_invoice(asset_id: str, current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)):
    try:
        object_id = ObjectId(asset_id)
    except Exception as error:  # pragma: no cover
        raise HTTPException(status_code=400, detail="Invalid asset id") from error

    item = await db["assets"].find_one({"_id": object_id, "user_id": current_user["id"]})
    if not item:
        raise HTTPException(status_code=404, detail="Asset not found")

    invoice_path = item.get("invoice_attachment_path")
    if not invoice_path:
        raise HTTPException(status_code=404, detail="Invoice not available")

    path = Path(str(invoice_path))
    if not path.exists():
        raise HTTPException(status_code=404, detail="Invoice file not found")

    return FileResponse(path, media_type="application/pdf", filename=path.name)


@router.post("/{asset_id}/documents")
async def upload_documents(asset_id: str, files: list[UploadFile] = File(...), current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)) -> dict[str, Any]:
    try:
        object_id = ObjectId(asset_id)
    except Exception as error:  # pragma: no cover
        raise HTTPException(status_code=400, detail="Invalid asset id") from error

    asset = await db["assets"].find_one({"_id": object_id, "user_id": current_user["id"]})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    saved_docs: list[dict[str, Any]] = []
    for upload in files:
        safe_name = upload.filename or "document"
        file_name = f"{asset_id}_{datetime.now(timezone.utc).timestamp()}_{safe_name}"
        file_path = DOC_ROOT / file_name
        content = await upload.read()
        file_path.write_bytes(content)

        doc = {
            "asset_id": asset_id,
            "user_id": current_user["id"],
            "file_name": safe_name,
            "document_type": "supporting",
            "file_path": str(file_path),
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
        }
        result = await db["asset_documents"].insert_one(doc)
        doc["_id"] = result.inserted_id
        saved_docs.append(_to_document(doc))

    return {"asset_id": asset_id, "uploaded": saved_docs}


@router.get("/{asset_id}/documents")
async def list_documents(asset_id: str, current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)) -> list[dict[str, Any]]:
    docs = await db["asset_documents"].find(
        {
            "asset_id": asset_id,
            "user_id": current_user["id"],
            "document_type": {"$in": ["supporting", "invoice"]},
        }
    ).sort("uploaded_at", -1).to_list(length=300)
    return [_to_document(doc) for doc in docs]


@router.get("/{asset_id}/documents/{document_id}/file")
async def get_document_file(asset_id: str, document_id: str, current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)):
    try:
        object_id = ObjectId(document_id)
    except Exception as error:  # pragma: no cover
        raise HTTPException(status_code=400, detail="Invalid document id") from error

    doc = await db["asset_documents"].find_one({"_id": object_id, "asset_id": asset_id, "user_id": current_user["id"]})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    path = Path(str(doc.get("file_path", "")))
    if not path.exists():
        raise HTTPException(status_code=404, detail="Document file not found")

    return FileResponse(path, media_type="application/octet-stream", filename=doc.get("file_name") or path.name)


@router.delete("/{asset_id}/documents/{document_id}")
async def delete_document(asset_id: str, document_id: str, current_user: dict[str, str] = Depends(get_current_user), db=Depends(get_db)) -> dict[str, str]:
    try:
        object_id = ObjectId(document_id)
    except Exception as error:  # pragma: no cover
        raise HTTPException(status_code=400, detail="Invalid document id") from error

    doc = await db["asset_documents"].find_one({"_id": object_id, "asset_id": asset_id, "user_id": current_user["id"]})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    path = Path(str(doc.get("file_path", "")))
    if path.exists():
        path.unlink(missing_ok=True)

    await db["asset_documents"].delete_one({"_id": object_id})
    return {"status": "deleted", "document_id": document_id}
